from PIL import Image, ImageDraw,   ImageFont,ImageFilter
import code128
from openpyxl import load_workbook
import os

def create(num,name,time):
    code128.image(num).save(f"qr_{num}.png")
    im1 = Image.open('main2.png')
    im3 = Image.open(f'qr_{num}.png')
    im2 = im3.resize((350,100))
    back_im = im1.copy()
    back_im.paste(im2, (1596, 409))
    back_im.save(f'with_qr_{num}.png', quality=95)
    width, height = 3600, 320
    width1, height1 = 3600, 450
    width2, height2 = 3600, 550


    color = 'rgb(0, 0, 0)'
    color2 = 'rgb(43, 35, 84)'

    username_font = ImageFont.truetype('Quicksand-VariableFont_wght.ttf', size=50)
    time_font = ImageFont.truetype('BrickSans-Bold.otf', size=55)
    (username_x, username_y) = (width/2-30, height/2+100)
    (username_x1, username_y1) = (width1/2-30, height1/2+100)
    (username_x2, username_y2) = (width2/2-30, height2/2+100)


    image = Image.open(f'with_qr_{num}.png')
    draw = ImageDraw.Draw(image)
                # campus_name = f"of branch {str.upper(df['Branch'][i])}, {str.lower(df['Year'][i])} Year (20{df['YearNo'][i]}-{df['YearNo'][i]+4}), {str.upper(df['Campus'][i])} who has actively taken participate"
    draw.text((username_x, username_y), name, fill=color, font=username_font,anchor="mm")
    draw.text((username_x1, username_y1), f"Report At {time}pm", fill=color2, font=time_font,anchor="mm")
    draw.text((username_x2, username_y2), f"SLOT {time}", fill=color2, font=time_font,anchor="mm")

                # draw.text((campus_x, campus_y), campus_name, fill=color, font=campus_font, anchor="mm")
    image.save(f'fguys/{num}.png')
    os.remove(f"qr_{num}.png")
    os.remove(f'with_qr_{num}.png')




wb = load_workbook("fguys.xlsx")
ws = wb.active
column = ws["J"]
number = [column[x].value for x in range(len(column))]
column2 = ws["C"]
name = [column2[x].value for x in range(len(column2))]
column3 = ws["K"]
time = [column3[x].value for x in range(len(column3))]

print(number)
for i in range(len(number)):
    if type(number[i]) == int:
        create(number[i],name[i].lower().capitalize(),time[i])

